#! /usr/bin/env python3
import pandas as pd
import sys
from openpyxl.styles import Alignment, Font

def extract_genus_species(taxonomic_string):
    parts = taxonomic_string.split(';')
    genus = None
    species = None
    for part in parts:
        if part.startswith('g__'):
            genus = part.split('__')[1]
        elif part.startswith('s__'):
            species = part.split('__')[1]
    return f"{genus} {species}".strip()

def calculate_abundance(input_file, classification_file, output_file):
    # Load the data
    data = pd.read_csv(input_file, sep='\s+')
    
    # Extract bin and contig names based on the delimiter "_k"
    data['Bin_Name'] = data['#Template'].apply(lambda x: x.rsplit('_k', 1)[0].strip())
    
    # Calculate the product of depth and template length for weighted coverage calculation
    data['Weighted_Coverage'] = data['Depth'] * data['Template_length']
    
    # Group by bin and calculate the weighted average coverage
    bin_grouped = data.groupby('Bin_Name').agg({
        'Weighted_Coverage': 'sum',
        'Template_length': 'sum'
    })
    bin_grouped['Abundance'] = bin_grouped['Weighted_Coverage'] / bin_grouped['Template_length']

    # Load the GTDB-Tk classification file
    classification_data = pd.read_csv(classification_file, sep='\t', header=0)
    classification_data['user_genome'] = classification_data['user_genome'].astype(str).str.strip()
    classification_data.set_index('user_genome', inplace=True)
    
    # Map the classification to the bins and convert to string
    bin_grouped['Taxonomic Classification'] = bin_grouped.index.map(classification_data['classification']).astype(str)
    bin_grouped['Genus and Species'] = bin_grouped['Taxonomic Classification'].apply(extract_genus_species)

    # Export to Excel with formatting adjustments
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        bin_grouped[['Abundance', 'Taxonomic Classification', 'Genus and Species']].to_excel(writer, sheet_name='Abundance Data', index_label='Bin Name')
        
        # Get the workbook and the worksheet
        workbook = writer.book
        worksheet = writer.sheets['Abundance Data']
        
        # Apply left alignment and non-bold font to all cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='left')
                cell.font = Font(bold=False)
                
        # Bold headers
        for cell in worksheet["1:1"]:  # This accesses the first row which is the header row
            cell.font = Font(bold=True)

if __name__ == "__main__":
    if len(sys.argv) != 4:
        print("Usage: ./calculate_abundance_bins.py <input_file.res> <classification_file.tsv> <output_file.xlsx>")
        sys.exit(1)
    
    input_file_path = sys.argv[1]
    classification_file_path = sys.argv[2]
    output_file_path = sys.argv[3]
    calculate_abundance(input_file_path, classification_file_path, output_file_path)
